from __future__ import annotations

from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _join_lines(items: List[str]) -> str:
    return "\n".join([i for i in items if (i or "").strip()])


def write_testcases_xlsx(payload: Dict[str, Any], out_path: str) -> None:
    wb = Workbook()

    # Sheet 1: TestCases
    ws = wb.active
    ws.title = "TestCases"

    headers = [
        "ID",
        "Category",
        "Priority",
        "Title",
        "Story Refs",
        "Preconditions",
        "Test Data",
        "Steps",
        "Expected Results",
        "Notes",
    ]
    ws.append(headers)

    for tc in payload.get("test_cases", []):
        ws.append([
            tc.get("id", ""),
            tc.get("category", ""),
            tc.get("priority", ""),
            tc.get("title", ""),
            ", ".join(tc.get("story_refs", []) or []),
            _join_lines(tc.get("preconditions", []) or []),
            _join_lines(tc.get("test_data", []) or []),
            _join_lines(tc.get("steps", []) or []),
            _join_lines(tc.get("expected_results", []) or []),
            _join_lines(tc.get("notes", []) or []),
        ])

    # Auto-width (simple heuristic)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 22

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Sheet 2: Gherkin
    ws2 = wb.create_sheet("Gherkin")
    headers2 = ["TC ID", "Feature", "Scenario", "Given", "When", "Then"]
    ws2.append(headers2)

    for tc in payload.get("test_cases", []):
        tc_id = tc.get("id", "")
        for g in (tc.get("gherkin", []) or []):
            ws2.append([
                tc_id,
                g.get("feature", ""),
                g.get("scenario", ""),
                _join_lines(g.get("given", []) or []),
                _join_lines(g.get("when", []) or []),
                _join_lines(g.get("then", []) or []),
            ])

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 24
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"

    wb.save(out_path)
